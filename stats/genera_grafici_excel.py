"""
Genera un file Excel con grafici delle statistiche risposte dei giocatori del Quizzone.
Include: percentuali corrette per quiz, punteggi, tempi medi, e overall.
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# DATI
# ============================================================

# Dati per quiz di ogni giocatore
players_data = {
    "Tato": {
        "quiz": ["P1", "P2", "P3"],
        "corrette": [20, 28, 31],
        "totali": [30, 30, 35],
        "percentuali": [67, 93, 89],
        "punteggi": [2879, 7072, 7554],
        "tempi": [7.7, 3.6, 4.3],
    },
    "Jacopo": {
        "quiz": ["P1", "P3"],
        "corrette": [21, 21],
        "totali": [30, 35],
        "percentuali": [70, 60],
        "punteggi": [2647, 2692],
        "tempi": [6.6, 6.5],
    },
    "Mattia": {
        "quiz": ["P1", "P2", "P3", "P5"],
        "corrette": [16, 18, 22, 24],
        "totali": [30, 30, 35, 35],
        "percentuali": [53, 60, 63, 69],
        "punteggi": [1754, 2917, 3181, 5014],
        "tempi": [6.1, 7.6, 6.8, 4.5],
    },
    "Manuel": {
        "quiz": ["P1", "P2"],
        "corrette": [13, 17],
        "totali": [30, 30],
        "percentuali": [43, 57],
        "punteggi": [1232, 2551],
        "tempi": [9.3, 7.6],
    },
}

# Overall
overall = {
    "Tato":   {"quiz_giocati": 3, "corrette_tot": 79, "domande_tot": 95, "perc": 83, "punti_tot": 17505, "tempo_medio": 5.2},
    "Jacopo": {"quiz_giocati": 2, "corrette_tot": 42, "domande_tot": 65, "perc": 65, "punti_tot": 5339, "tempo_medio": 6.6},
    "Mattia": {"quiz_giocati": 4, "corrette_tot": 80, "domande_tot": 130, "perc": 62, "punti_tot": 12866, "tempo_medio": 6.3},
    "Manuel": {"quiz_giocati": 2, "corrette_tot": 30, "domande_tot": 60, "perc": 50, "punti_tot": 3783, "tempo_medio": 8.5},
}

# ============================================================
# FOGLIO 1: Overall
# ============================================================
ws1 = wb.active
ws1.title = "Overall"

# Intestazioni
headers = ["Giocatore", "Quiz giocati", "Corrette", "Domande tot", "% Corrette", "Punti totali", "Tempo medio (s)"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

row = 2
for name, data in overall.items():
    ws1.cell(row=row, column=1, value=name).border = thin_border
    ws1.cell(row=row, column=2, value=data["quiz_giocati"]).border = thin_border
    ws1.cell(row=row, column=3, value=data["corrette_tot"]).border = thin_border
    ws1.cell(row=row, column=4, value=data["domande_tot"]).border = thin_border
    ws1.cell(row=row, column=5, value=data["perc"]).border = thin_border
    ws1.cell(row=row, column=6, value=data["punti_tot"]).border = thin_border
    ws1.cell(row=row, column=7, value=data["tempo_medio"]).border = thin_border
    for col in range(1, 8):
        ws1.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    row += 1

# Adatta larghezza colonne
for col in range(1, 8):
    ws1.column_dimensions[get_column_letter(col)].width = 16

# Grafico barre: % Corrette overall
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Percentuale Corrette — Overall"
chart1.y_axis.title = "% Corrette"
chart1.x_axis.title = "Giocatore"
chart1.style = 10

data_ref = Reference(ws1, min_col=5, min_row=1, max_row=5)
cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=5)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
chart1.width = 14
chart1.height = 10
ws1.add_chart(chart1, "A8")

# Grafico barre: Punteggio totale
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Punteggio Totale"
chart2.y_axis.title = "Punti"
chart2.x_axis.title = "Giocatore"
chart2.style = 10

data_ref2 = Reference(ws1, min_col=6, min_row=1, max_row=5)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.width = 14
chart2.height = 10
ws1.add_chart(chart2, "I8")

# Grafico barre: Tempo medio
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Tempo Medio di Risposta"
chart3.y_axis.title = "Secondi"
chart3.x_axis.title = "Giocatore"
chart3.style = 10

data_ref3 = Reference(ws1, min_col=7, min_row=1, max_row=5)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref)
chart3.width = 14
chart3.height = 10
ws1.add_chart(chart3, "A25")

# ============================================================
# FOGLIO 2: Progressione per quiz
# ============================================================
ws2 = wb.create_sheet("Progressione")

# Tabella percentuali per quiz
ws2.cell(row=1, column=1, value="Quiz").font = Font(bold=True)
quiz_labels = ["P1", "P2", "P3", "P5"]
for i, q in enumerate(quiz_labels, 2):
    ws2.cell(row=1, column=i, value=q).font = Font(bold=True)
    ws2.cell(row=1, column=i).alignment = Alignment(horizontal="center")

row = 2
for name, pdata in players_data.items():
    ws2.cell(row=row, column=1, value=name)
    for q, perc in zip(pdata["quiz"], pdata["percentuali"]):
        col_idx = quiz_labels.index(q) + 2
        ws2.cell(row=row, column=col_idx, value=perc)
    row += 1

# Grafico linee: progressione % corrette
chart4 = LineChart()
chart4.title = "Progressione % Corrette per Puntata"
chart4.y_axis.title = "% Corrette"
chart4.x_axis.title = "Puntata"
chart4.style = 10
chart4.width = 18
chart4.height = 12

data_ref4 = Reference(ws2, min_col=2, max_col=5, min_row=1, max_row=5)
cats_ref4 = Reference(ws2, min_col=2, max_col=5, min_row=1)
chart4.add_data(data_ref4, from_rows=False, titles_from_data=True)

# Categorie dai quiz labels
cats_ws2 = Reference(ws2, min_col=1, min_row=2, max_row=5)
# Usiamo approccio diverso: dati per righe
# Riscriviamo la tabella in modo più adatto ai grafici a linee

# Riscrittura: righe = quiz, colonne = giocatori
ws2.delete_rows(1, 10)

ws2.cell(row=1, column=1, value="Puntata").font = Font(bold=True)
player_names = ["Tato", "Jacopo", "Mattia", "Manuel"]
for i, name in enumerate(player_names, 2):
    ws2.cell(row=1, column=i, value=name).font = Font(bold=True)
    ws2.cell(row=1, column=i).alignment = Alignment(horizontal="center")

# Percentuali per puntata
quiz_all = ["P1", "P2", "P3", "P5"]
perc_matrix = {
    "P1": {"Tato": 67, "Jacopo": 70, "Mattia": 53, "Manuel": 43},
    "P2": {"Tato": 93, "Jacopo": None, "Mattia": 60, "Manuel": 57},
    "P3": {"Tato": 89, "Jacopo": 60, "Mattia": 63, "Manuel": None},
    "P5": {"Tato": None, "Jacopo": None, "Mattia": 69, "Manuel": None},
}

for r, quiz in enumerate(quiz_all, 2):
    ws2.cell(row=r, column=1, value=quiz)
    for c, name in enumerate(player_names, 2):
        val = perc_matrix[quiz][name]
        if val is not None:
            ws2.cell(row=r, column=c, value=val)

# Grafico linee
chart4 = LineChart()
chart4.title = "Progressione % Corrette per Puntata"
chart4.y_axis.title = "% Corrette"
chart4.x_axis.title = "Puntata"
chart4.style = 10
chart4.width = 18
chart4.height = 12
chart4.y_axis.scaling.min = 30
chart4.y_axis.scaling.max = 100

data_ref4 = Reference(ws2, min_col=2, max_col=5, min_row=1, max_row=5)
cats_ref4 = Reference(ws2, min_col=1, min_row=2, max_row=5)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
ws2.add_chart(chart4, "A8")

# Tabella punteggi per puntata
ws2.cell(row=1, column=8, value="Puntata").font = Font(bold=True)
for i, name in enumerate(player_names, 9):
    ws2.cell(row=1, column=i, value=name + " (pt)").font = Font(bold=True)

punti_matrix = {
    "P1": {"Tato": 2879, "Jacopo": 2647, "Mattia": 1754, "Manuel": 1232},
    "P2": {"Tato": 7072, "Jacopo": None, "Mattia": 2917, "Manuel": 2551},
    "P3": {"Tato": 7554, "Jacopo": 2692, "Mattia": 3181, "Manuel": None},
    "P5": {"Tato": None, "Jacopo": None, "Mattia": 5014, "Manuel": None},
}

for r, quiz in enumerate(quiz_all, 2):
    ws2.cell(row=r, column=8, value=quiz)
    for c, name in enumerate(player_names, 9):
        val = punti_matrix[quiz][name]
        if val is not None:
            ws2.cell(row=r, column=c, value=val)

# Grafico punteggi
chart5 = BarChart()
chart5.type = "col"
chart5.title = "Punteggio per Puntata"
chart5.y_axis.title = "Punti"
chart5.x_axis.title = "Puntata"
chart5.style = 10
chart5.width = 18
chart5.height = 12

data_ref5 = Reference(ws2, min_col=9, max_col=12, min_row=1, max_row=5)
cats_ref5 = Reference(ws2, min_col=8, min_row=2, max_row=5)
chart5.add_data(data_ref5, titles_from_data=True)
chart5.set_categories(cats_ref5)
ws2.add_chart(chart5, "A25")

# ============================================================
# FOGLIO 3: Tempi medi
# ============================================================
ws3 = wb.create_sheet("Tempi")

ws3.cell(row=1, column=1, value="Puntata").font = Font(bold=True)
for i, name in enumerate(player_names, 2):
    ws3.cell(row=1, column=i, value=name).font = Font(bold=True)

tempi_matrix = {
    "P1": {"Tato": 7.7, "Jacopo": 6.6, "Mattia": 6.1, "Manuel": 9.3},
    "P2": {"Tato": 3.6, "Jacopo": None, "Mattia": 7.6, "Manuel": 7.6},
    "P3": {"Tato": 4.3, "Jacopo": 6.5, "Mattia": 6.8, "Manuel": None},
    "P5": {"Tato": None, "Jacopo": None, "Mattia": 4.5, "Manuel": None},
}

for r, quiz in enumerate(quiz_all, 2):
    ws3.cell(row=r, column=1, value=quiz)
    for c, name in enumerate(player_names, 2):
        val = tempi_matrix[quiz][name]
        if val is not None:
            ws3.cell(row=r, column=c, value=val)

# Grafico tempi
chart6 = LineChart()
chart6.title = "Tempo Medio di Risposta per Puntata"
chart6.y_axis.title = "Secondi"
chart6.x_axis.title = "Puntata"
chart6.style = 10
chart6.width = 18
chart6.height = 12

data_ref6 = Reference(ws3, min_col=2, max_col=5, min_row=1, max_row=5)
cats_ref6 = Reference(ws3, min_col=1, min_row=2, max_row=5)
chart6.add_data(data_ref6, titles_from_data=True)
chart6.set_categories(cats_ref6)
ws3.add_chart(chart6, "A8")

# ============================================================
# FOGLIO 4: Dettaglio risposte (corrette vs sbagliate)
# ============================================================
ws4 = wb.create_sheet("Dettaglio Risposte")

ws4.cell(row=1, column=1, value="Giocatore").font = Font(bold=True)
ws4.cell(row=1, column=2, value="Corrette totali").font = Font(bold=True)
ws4.cell(row=1, column=3, value="Sbagliate totali").font = Font(bold=True)

row = 2
for name, data in overall.items():
    ws4.cell(row=row, column=1, value=name)
    ws4.cell(row=row, column=2, value=data["corrette_tot"])
    ws4.cell(row=row, column=3, value=data["domande_tot"] - data["corrette_tot"])
    row += 1

# Grafico barre impilate corrette/sbagliate
chart7 = BarChart()
chart7.type = "col"
chart7.grouping = "stacked"
chart7.title = "Risposte Corrette vs Sbagliate — Totale"
chart7.y_axis.title = "Numero risposte"
chart7.style = 10
chart7.width = 16
chart7.height = 12

data_ref7 = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=5)
cats_ref7 = Reference(ws4, min_col=1, min_row=2, max_row=5)
chart7.add_data(data_ref7, titles_from_data=True)
chart7.set_categories(cats_ref7)

# Colori: verde per corrette, rosso per sbagliate
chart7.series[0].graphicalProperties.solidFill = "70AD47"  # verde
chart7.series[1].graphicalProperties.solidFill = "FF4444"  # rosso

ws4.add_chart(chart7, "A8")

# Adatta colonne
for ws in [ws2, ws3, ws4]:
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14

# ============================================================
# SALVA
# ============================================================
output_path = r"c:\Users\davidea\Saved Games\Skars-1-taking-the-finals\stats\quizzone_statistiche.xlsx"
wb.save(output_path)
print(f"File salvato: {output_path}")
